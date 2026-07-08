#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment
from pathlib import Path

ROOT = Path("/home/user/workspace/legal-work/target/testakten/statusfeststellung-gmbh-geschaeftsfuehrer-minderheit-erlangen")

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()

# Sheet 1: Rentenversicherung + Arbeitslosenversicherung Beitragsberechnung je Monat/Jahr
ws1 = wb.active
ws1.title = "RV-ALV-Beitraege"
ws1.append(["Zeitraum", "Bruttoverguetung EUR/Monat", "Beitragsbemessungsgrenze RV/Monat EUR", "Beitragspflichtiges Entgelt EUR/Monat", "RV-Beitragssatz Prozent", "ALV-Beitragssatz Prozent", "RV-Beitrag AG+AN EUR/Monat", "ALV-Beitrag AG+AN EUR/Monat"])
rows1 = [
    ["2023-03 bis 2023-12", 12000, 7300, 7300, "18.6", "2.6", 1357.80, 189.80],
    ["2024-01 bis 2024-12", 12000, 7550, 7550, "18.6", "2.6", 1404.30, 196.30],
    ["2025-01 bis 2025-12", 12000, 7550, 7550, "18.6", "2.6", 1404.30, 196.30],
    ["2026-01 bis 2026-06", 12000, 7450, 7450, "18.6", "2.6", 1385.70, 193.70],
]
for r in rows1:
    ws1.append(r)
style_header(ws1, 8)
ws1.append([])
ws1.append(["Hinweis"])
ws1.append(["Die Beitragsbemessungsgrenzen sind bundeseinheitliche Werte laut Sozialversicherungs-Rechengroessenverordnung fuer die jeweiligen Jahre. Fuer 2026 werden vorlaeufige Werte verwendet, endgueltige Werte stehen zum Zeitpunkt der Aktenanlage noch aus."])

# Sheet 2: Nachforderungsaufstellung mit Verjaehrung
ws2 = wb.create_sheet("Nachforderung-Paragraf25")
ws2.append(["Zeitraum", "RV-Beitrag Summe EUR", "ALV-Beitrag Summe EUR", "Summe EUR", "Verjaehrungsfrist Paragraf 25 SGB IV", "Verjaehrung tritt ein zum", "Bemerkung"])
rows2 = [
    ["2023-03 bis 2023-12 (10 Monate)", 13578.00, 1898.00, 15476.00, "4 Jahre (Regelfrist)", "31.12.2027", "Nicht verjaehrt, sofern kein Vorsatz"],
    ["2024-01 bis 2024-12 (12 Monate)", 16851.60, 2355.60, 19207.20, "4 Jahre (Regelfrist)", "31.12.2028", "Nicht verjaehrt"],
    ["2025-01 bis 2025-12 (12 Monate)", 16851.60, 2355.60, 19207.20, "4 Jahre (Regelfrist)", "31.12.2029", "Nicht verjaehrt"],
    ["2026-01 bis 2026-06 (6 Monate)", 8314.20, 1162.20, 9476.40, "4 Jahre (Regelfrist)", "31.12.2030", "Nicht verjaehrt, Verfahren laeuft"],
]
for r in rows2:
    ws2.append(r)
style_header(ws2, 7)
ws2.append([])
ws2.append(["Gesamtsumme Hauptforderung", "", "", "=SUM(D2:D5)"])
ws2.append([])
ws2.append(["Pruefhinweis"])
ws2.append(["Die vierjaehrige Regelverjaehrung nach Paragraf 25 Absatz 1 SGB IV gilt, solange kein bedingter Vorsatz vorliegt; bei Vorsatz betruege die Frist 30 Jahre. Die Gesellschaft hatte sich auf eine rechtliche Einschaetzung ihres damaligen Beraters gestuetzt, sodass Vorsatz von der Kanzlei bestritten wird."])

# Sheet 3: Saeumniszuschlaege
ws3 = wb.create_sheet("Saeumniszuschlaege")
ws3.append(["Zeitraum", "Rueckstand EUR (gerundet)", "Saeumniszuschlag Prozent pro Monat", "Berechnungsbeginn", "Bemerkung"])
rows3 = [
    ["2023-03 bis 2025-11", 15476.00, "1.0", "erst ab Kenntnis, siehe Anhoerung 18.11.2025", "Gutglaeubigkeit wird von der Kanzlei geltend gemacht"],
    ["2025-12 bis laufend", 19207.20, "1.0", "ab 01.12.2025", "unstreitig, falls Bescheid bestandskraeftig wird"],
]
for r in rows3:
    ws3.append(r)
style_header(ws3, 5)

wb.save(ROOT / "31_beitragsberechnung_rv_kv_pv_alv.xlsx")
print("geschrieben:", ROOT / "31_beitragsberechnung_rv_kv_pv_alv.xlsx")
