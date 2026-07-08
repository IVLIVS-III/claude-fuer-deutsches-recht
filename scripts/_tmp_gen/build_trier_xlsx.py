from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

BASE = "/home/user/workspace/legal-work/target/testakten/unfallversicherung-arbeitsunfall-lagerleiter-sturz-trier/"

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

wb = Workbook()

# Sheet 1: Verletztengeldberechnung mit AAW-Vergleich
ws1 = wb.active
ws1.title = "Verletztengeld-AAW"
ws1.append(["Verletztengeldberechnung mit Vergleich zum Regelentgelt (AAW)"])
ws1["A1"].font = Font(bold=True, size=13)
ws1.append([])
ws1.append(["Monat", "Bruttoarbeitsentgelt", "Regelentgelt kalendertaeglich", "Verletztengeld 80 Prozent Brutto", "Verletztengeld kalendertaeglich", "Nettoarbeitsentgelt Vergleich", "Bemerkung"])
style_header(ws1, 3, 7)
rows = [
    ["April 2026 (bis 03.04.)", 4180, 139.33, 111.46, 111.46, 128.50, "Entgeltfortzahlung bis 03.04., danach Verletztengeld ab 15.05."],
    ["Mai 2026", 4180, 139.33, 111.46, 111.46, 128.50, "Entgeltfortzahlung 6 Wochen bis 14.05., danach Verletztengeld"],
    ["Juni 2026", 4180, 139.33, 111.46, 111.46, 128.50, "Verletztengeld streitig wegen Ablehnungsbescheid"],
    ["Juli 2026", 4180, 139.33, 111.46, 111.46, 128.50, "Widerspruchsmonat, Zahlung ausgesetzt"],
    ["August 2026", 4180, 139.33, 111.46, 111.46, 128.50, "Weiterhin ausgesetzt bis Widerspruchsbescheid"],
]
for r in rows:
    ws1.append(r)
ws1.append([])
ws1.append(["Hinweis: Verletztengeld betraegt 80 Prozent des Regelentgelts, gedeckelt auf das Nettoarbeitsentgelt (Paragraf 47 SGB VII). Die Berechnung ist vorlaeufig und wurde mangels Anerkennung als Arbeitsunfall bislang nicht ausgezahlt."])
for col, width in zip("ABCDEFG", [26, 20, 24, 24, 22, 24, 46]):
    ws1.column_dimensions[col].width = width

# Sheet 2: MdE-Berechnung
ws2 = wb.create_sheet("MdE-Berechnung")
ws2.append(["Einschaetzung der Minderung der Erwerbsfaehigkeit (MdE)"])
ws2["A1"].font = Font(bold=True, size=13)
ws2.append([])
ws2.append(["Funktionsbereich", "Befund", "Vergleichswert unfallmedizinische Tabelle", "Einzel-MdE"])
style_header(ws2, 3, 4)
rows2 = [
    ["Aktive Abduktion rechts", "150 Grad, Kraftgrad 4 von 5", "Verlust der Armhebefaehigkeit ueber Schulterhoehe bei sonst freier Beweglichkeit 20 bis 30 vom Hundert", "20 vom Hundert"],
    ["Aussenrotation", "endgradig schmerzhaft eingeschraenkt", "geringe zusaetzliche Funktionsminderung", "kein Zuschlag, in Gesamtwert eingerechnet"],
    ["Kraftminderung Gesamtarm", "leicht reduziert im Seitenvergleich", "keine eigenstaendige Erhoehung", "kein Zuschlag"],
]
for r in rows2:
    ws2.append(r)
ws2.append([])
ws2.append(["Gutachterliche Gesamt-MdE (vorlaeufig)", "", "", "20 vom Hundert"])
ws2.append([])
ws2.append(["Hinweis: Eine MdE von mindestens 20 vom Hundert ueber die 26. Woche nach dem Arbeitsunfall hinaus begruendet dem Grunde nach einen Anspruch auf Verletztenrente nach Paragraf 56 SGB VII. Eine Nachuntersuchung nach zwoelf Monaten wurde vom Sachverstaendigen empfohlen."])
for col, width in zip("ABCD", [30, 34, 55, 20]):
    ws2.column_dimensions[col].width = width

# Sheet 3: Rentenberechnung Unfallrente
ws3 = wb.create_sheet("Rentenberechnung-Unfallrente")
ws3.append(["Vorlaeufige Berechnung der Verletztenrente"])
ws3["A1"].font = Font(bold=True, size=13)
ws3.append([])
ws3.append(["Berechnungsgroesse", "Wert", "Erlaeuterung"])
style_header(ws3, 3, 3)
rows3 = [
    ["Jahresarbeitsverdienst (JAV)", "50160 Euro", "Bruttojahresverdienst der zwoelf Monate vor dem Unfall, 12 mal 4180 Euro"],
    ["MdE", "20 vom Hundert", "Vorlaeufige gutachterliche Einschaetzung"],
    ["Vollrente (bei MdE 100)", "33440 Euro pro Jahr", "Zwei Drittel des JAV nach Paragraf 56 Absatz 3 SGB VII"],
    ["Teilrente bei MdE 20", "6688 Euro pro Jahr", "20 vom Hundert der Vollrente"],
    ["Monatliche Teilrente", "557.33 Euro", "Teilrente geteilt durch zwoelf Monate"],
]
for r in rows3:
    ws3.append(r)
ws3.append([])
ws3.append(["Hinweis: Die Berechnung ist eine rentenberaterliche Modellrechnung auf Grundlage der vorlaeufigen MdE-Einschaetzung von 20 vom Hundert und dient der Vorbereitung des Klageverfahrens. Endgueltig entscheidet die Berufsgenossenschaft beziehungsweise das Sozialgericht nach Vorliegen der Nachuntersuchung."])
for col, width in zip("ABC", [30, 22, 55]):
    ws3.column_dimensions[col].width = width

wb.save(BASE + "30_verletztengeld_mde_rentenberechnung.xlsx")
print("XLSX geschrieben")
