import openpyxl
from openpyxl.styles import Font, Alignment
import os

BASE = "/home/user/workspace/legal-work/target/testakten/sozialrecht-elektrorollstuhl-koerner-oldenburg/xlsx"
os.makedirs(BASE, exist_ok=True)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = "Verfahrenskalender"
ws1.append(["Datum", "Ereignis", "Frist", "Zustaendig"])
rows1 = [
    ["12.05.2026", "Antrag Elektrorollstuhl gestellt", "", "Kläger"],
    ["07.05.2026", "Aerztliche Verordnung Dr. Stahlmann", "", "Hausarzt"],
    ["17.05.2026", "Facharztbericht Dr. Radtke", "", "Facharzt"],
    ["20.05.2026", "Kostenvoranschlag RehaTechnik Albrecht", "", "Sanitaetshaus"],
    ["05.06.2026", "MD-Stellungnahme Ablehnungsempfehlung", "", "Medizinischer Dienst"],
    ["10.06.2026", "Ablehnungsbescheid", "", "Kasse"],
    ["18.06.2026", "Privates Reha-Gutachten", "", "RehaKompetenz Nordwest"],
    ["20.06.2026", "Widerspruch eingelegt", "fristgerecht", "Kläger"],
    ["24.06.2026", "Widerspruchsbescheid", "", "Kasse"],
    ["26.06.2026", "Zustellung Widerspruchsbescheid", "Klagefrist beginnt", "Kläger"],
    ["29.06.2026", "Mandatierung Kanzlei Brammertz", "", "Kläger"],
    ["03.07.2026", "Klageschrift gefertigt", "", "Kanzlei"],
    ["27.07.2026", "Ende Klagefrist", "letzter Tag", "Kanzlei"],
    ["22.07.2026", "Klageerwiderung Kasse", "", "Kasse"],
    ["05.08.2026", "Beweisbeschluss Sozialgericht", "", "Sozialgericht Oldenburg"],
    ["02.09.2026", "Untersuchungstermin Sachverstaendige", "", "Prof. Dr. Wickenrath"],
    ["16.09.2026", "Gutachten vorgelegt", "", "Prof. Dr. Wickenrath"],
    ["25.09.2026", "Terminsverfuegung", "", "Sozialgericht Oldenburg"],
    ["22.10.2026", "Muendliche Verhandlung", "10.30 Uhr, Saal 3", "Sozialgericht Oldenburg"],
]
for r in rows1:
    ws1.append(r)
style_header(ws1)
for col in ["A", "B", "C", "D"]:
    ws1.column_dimensions[col].width = 30

ws2 = wb.create_sheet("Kostenvergleich")
ws2.append(["Hilfsmittel", "Kosten in EUR", "Foerderfaehig", "Bemerkung"])
rows2 = [
    ["Unterarmgehstuetze", 45.00, "ja", "bereits vorhanden"],
    ["Rollator", 0.00, "ja", "bereits vorhanden"],
    ["Standard-Faltrollstuhl", 740.00, "ja", "bereits vorhanden"],
    ["Elektromobil", 4200.00, "strittig", "Wendekreis und Abstellproblem im Altbau"],
    ["Elektrorollstuhl Basis", 8990.00, "beantragt", "Hauptangebot Kostenvoranschlag"],
    ["Elektrorollstuhl mit Sitzlift", 10990.00, "beantragt Alternative", "Zusatzausstattung"],
    ["Elektrorollstuhl Standardmodell Kasse", 6480.00, "Gegenangebot Kasse", "geringere Reichweite und Zuladung"],
]
for r in rows2:
    ws2.append(r)
style_header(ws2)
for col in ["A", "B", "C", "D"]:
    ws2.column_dimensions[col].width = 28

wb.save(f"{BASE}/verfahrenskalender_und_kostenvergleich.xlsx")
print("geschrieben: verfahrenskalender_und_kostenvergleich.xlsx")
