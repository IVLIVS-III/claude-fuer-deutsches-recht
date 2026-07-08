from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE = "/home/user/workspace/legal-work/target/testakten/unfallversicherung-arbeitsunfall-lagerleiter-sturz-trier/"

wb = Workbook()
ws = wb.active
ws.title = "Fehlzeiten-Uebersicht"
ws.append(["Fehlzeitenuebersicht Karsten Moeller, 2016 bis 2026"])
ws["A1"].font = Font(bold=True, size=13)
ws.append([])
ws.append(["Jahr", "Fehltage gesamt", "davon schulterbezogen", "Bemerkung"])
for c in range(1, 5):
    cell = ws.cell(row=3, column=c)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
rows = [
    [2016, 3, 2, "Schulterbeschwerden nach Handball, kurzfristig"],
    [2017, 2, 2, "Erneute kurze Schonung nach Sport"],
    [2018, 1, 1, "Letzte schulterbezogene Fehlzeit vor dem Unfall"],
    [2019, 4, 0, "Grippaler Infekt, keine Schulterbeschwerden"],
    [2020, 0, 0, "Keine Fehltage"],
    [2021, 2, 0, "Grippaler Infekt"],
    [2022, 0, 0, "Keine Fehltage"],
    [2023, 1, 0, "Zahnbehandlung"],
    [2024, 0, 0, "Keine Fehltage, Vorsorgeuntersuchung uneingeschraenkt tauglich"],
    [2025, 0, 0, "Keine Fehltage"],
    [2026, 68, 68, "Ab 03.04.2026 durchgehend wegen Unfallereignis"],
]
for r in rows:
    ws.append(r)
ws.append([])
ws.append(["Hinweis: Zwischen 2019 und dem Unfalltag 2026 gab es keine einzige schulterbezogene Fehlzeit, was die betriebliche Beschwerdefreiheit in den letzten sieben Jahren vor dem Ereignis belegt."])
for col, width in zip("ABCD", [10, 18, 24, 50]):
    ws.column_dimensions[col].width = width

wb.save(BASE + "xlsx/fehlzeiten_uebersicht_2016_2026.xlsx")
print("geschrieben")
